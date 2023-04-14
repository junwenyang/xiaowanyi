import openpyxl as xl

xl_file = "dyson.xlsx"
xl_dyson = xl.load_workbook(xl_file, data_only = True)

# 公式页
sh_formula = xl_dyson['公式']
rows_formula = sh_formula.max_row
rows = sh_formula["A2":"D"+str(rows_formula)]
formula = []
for cell1, cell2, cell3, cell4 in rows:
    formula.append((cell1.value, cell2.value, cell3.value, cell4.value))

# 计算
def costCal(galaxy):
    sh_galaxy = xl_dyson[galaxy]
    rows_total = sh_galaxy.max_row
    rows = sh_galaxy["B3":"C"+str(rows_total)]
    galaxy = []
    for cell1, cell2 in rows:
        galaxy.append((cell1.value, cell2.value))

    tmp = {}
    for g in galaxy:
        for f in formula:
            if(f[1]==g[0] and f[1] is not None and g[0] is not None):
                for g1 in galaxy:
                    if(g1[0] == f[0]):
                        print(g1[1])
                        cost = f[2]*g1[1]
                        tmp[f[1]] = cost + int(tmp.get(f[1], 0.0))
    
    print("消耗：\n",tmp)

    for i in range(3, rows_total):
        for key in tmp:
            if(key == sh_galaxy.cell(column = 2, row = i).value):
                sh_galaxy.cell(column = 4, row = i, value = tmp.get(key))

costCal('产物汇总')
xl_dyson.save('dyson1.xlsx')
# 自动回写主文件（待）